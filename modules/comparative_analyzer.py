import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

class ComparativeAnalyzer:
    def __init__(self, logger):
        self.logger = logger
        self.rojo = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        self.verde = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
        self.fuente_negra = Font(color="000000")
        
    def procesar_consolidado(self, ruta_consolidado):
        try:
            self.logger.agregar_log(f"\nProcesando archivo consolidado: {ruta_consolidado}...")

            df = pd.read_excel(
                ruta_consolidado,
                sheet_name="Consolidado",
                dtype={'CODIGO_SIIGO': str}  # Forzar lectura como string
            )

            df.columns = df.columns.str.strip()
            df["REFERENCIA"] = df["REFERENCIA"].fillna("").astype(str).str.strip().replace("nan", "")
            
            # Modificación clave: Eliminar la transformación que agregaba ceros
            df["CODIGO_SIIGO"] = df["CODIGO_SIIGO"].fillna("").astype(str).str.strip().replace("nan", "")

            df_con_ref = df[df["REFERENCIA"] != ""].copy()
            df_sin_ref = df[df["REFERENCIA"] == ""].copy()

            columnas_finales = [
                "REFERENCIA", 
                "DESCRIPCION",
                "ORIGEN",
                "INVENTARIO MANUAL",
                "SIIGO",
                "DIFERENCIA",
                "UBICACION_MARCAS",
                "UBICACION_SIIGO",
                "CODIGO_SIIGO"
            ]

            if not df_con_ref.empty:
                agrupado = df_con_ref.groupby("REFERENCIA")
                registros = []
                for ref, grupo in agrupado:
                    grupo = grupo.copy()
                    
                    descripcion = grupo['DESCRIPCION'][grupo['DESCRIPCION'].notna() & (grupo['DESCRIPCION'].astype(str).str.strip() != "")].astype(str)
                    descripcion = descripcion.iloc[0] if not descripcion.empty else ""
                    
                    origenes = sorted(grupo['ORIGEN'].dropna().unique())
                    origen_str = ", ".join(origenes)
                    
                    inventario_manual = grupo[grupo["ORIGEN"].isin(["STIHL", "SUZUKI", "YAMAHA"])]["CANTIDAD"].sum()
                    siigo = grupo[grupo["ORIGEN"] == "SIIGO"]["CANTIDAD"].sum()
                    diferencia = inventario_manual - siigo

                    ubicacion_marcas = ", ".join(
                        sorted(set(
                            grupo[grupo["ORIGEN"].isin(["STIHL", "SUZUKI", "YAMAHA"])]["UBICACION"]
                            .dropna().astype(str).str.strip().replace("nan", "")
                        ))
                    )

                    ubicacion_siigo = ", ".join(
                        sorted(set(
                            grupo[grupo["ORIGEN"] == "SIIGO"]["UBICACION"]
                            .dropna().astype(str).str.strip().replace("nan", "")
                        ))
                    )

                    codigos_validos = []
                    for codigo in grupo["CODIGO_SIIGO"]:
                        codigo_str = str(codigo).strip()
                        if codigo_str and codigo_str.lower() != "nan":
                            codigos_validos.append(codigo_str)
                    codigos_siigo = ", ".join(sorted(set(codigos_validos)))

                    registros.append({
                        "REFERENCIA": ref,
                        "DESCRIPCION": descripcion,
                        "ORIGEN": origen_str,
                        "INVENTARIO MANUAL": inventario_manual,
                        "SIIGO": siigo,
                        "DIFERENCIA": diferencia,
                        "UBICACION_MARCAS": ubicacion_marcas,
                        "UBICACION_SIIGO": ubicacion_siigo,
                        "CODIGO_SIIGO": codigos_siigo
                    })

                tabla_final = pd.DataFrame(registros)
            else:
                tabla_final = pd.DataFrame(columns=columnas_finales)

            if not df_sin_ref.empty:
                sin_ref_processed = df_sin_ref.copy()

                sin_ref_processed['INVENTARIO MANUAL'] = 0
                sin_ref_processed['SIIGO'] = sin_ref_processed['CANTIDAD']
                sin_ref_processed['DIFERENCIA'] = sin_ref_processed['CANTIDAD']
                sin_ref_processed['UBICACION_MARCAS'] = sin_ref_processed['UBICACION'].apply(
                    lambda x: str(x).strip() if str(x).strip().lower() not in ["", "nan"] else ""
                )
                sin_ref_processed['UBICACION_SIIGO'] = ""
                sin_ref_processed['CODIGO_SIIGO'] = sin_ref_processed['CODIGO_SIIGO'].fillna('').astype(str)

                sin_ref_processed = sin_ref_processed[[
                    "REFERENCIA", "DESCRIPCION", "ORIGEN", "INVENTARIO MANUAL",
                    "SIIGO", "DIFERENCIA", "UBICACION_MARCAS", "UBICACION_SIIGO", "CODIGO_SIIGO"
                ]]
                
                tabla_final = pd.concat([tabla_final, sin_ref_processed], ignore_index=True)

            analisis_file = Path("outputs") / "Analisis_Comparativo.xlsx"
            with pd.ExcelWriter(analisis_file, engine='openpyxl') as writer:
                tabla_final.to_excel(writer, sheet_name='Comparativo', index=False)
                
                ws = writer.sheets['Comparativo']
                
                for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
                    for cell in row:
                        if cell.value > 0:
                            cell.fill = self.verde
                        elif cell.value < 0:
                            cell.fill = self.rojo
                        cell.font = self.fuente_negra

                for sheet in writer.sheets.values():
                    sheet.sheet_state = 'visible'

            self.logger.agregar_log(f"Análisis comparativo creado: {analisis_file}", 'exito')
            return analisis_file

        except Exception as e:
            self.logger.agregar_log(f"Error durante el procesamiento: {str(e)}", 'error')
            raise
