import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

class FileProcessor:
    def __init__(self, logger):
        self.logger = logger

    def _buscar_fila_encabezados(self, df_raw, columnas_clave):
        """Busca la fila que contiene los encabezados requeridos."""
        for i, fila in df_raw.iterrows():
            valores = [str(cell).upper().strip() for cell in fila.values]
            if all(any(col in valores for col in cols) for cols in columnas_clave):
                return i
        return None

    def _mapear_columnas(self, df, mapeo_columnas):
        """Renombra columnas según el mapeo proporcionado."""
        columnas_renombradas = {}
        for col in df.columns:
            col_upper = str(col).upper().strip()
            for key, values in mapeo_columnas.items():
                if any(v == col_upper for v in values):
                    columnas_renombradas[col] = key
                    break
            else:
                columnas_renombradas[col] = col
        return df.rename(columns=columnas_renombradas)

    def _procesar_archivo_generico(self, archivo, mapeo_columnas, columnas_requeridas, origen):
        """Procesamiento genérico para archivos de inventario."""
        try:
            self.logger.agregar_log(f"Procesando archivo {archivo.name}...")

            df_raw = pd.read_excel(archivo, header=None)
            columnas_clave = [[col for col in cols] for cols in mapeo_columnas.values()]
            fila_encabezados = self._buscar_fila_encabezados(df_raw, columnas_clave)

            if fila_encabezados is None:
                raise ValueError("No se encontró fila con los encabezados requeridos")

            df = pd.read_excel(archivo, header=fila_encabezados)
            df.columns = df.columns.str.upper().str.strip()
            df = self._mapear_columnas(df, mapeo_columnas)

            for col in columnas_requeridas:
                if col not in df.columns:
                    raise ValueError(f"Falta columna requerida: {col}")

            # Limpieza de datos
            df = df.dropna(how='all')
            if 'REFERENCIA' in df.columns:
                df['REFERENCIA'] = df['REFERENCIA'].astype(str).str.strip()
            if 'DESCRIPCION' in df.columns:
                df['DESCRIPCION'] = df['DESCRIPCION'].astype(str).str.strip()
            if 'CANTIDAD' in df.columns:
                df['CANTIDAD'] = pd.to_numeric(df['CANTIDAD'], errors='coerce').fillna(0)

            # Agregar columnas adicionales si no existen
            if 'UBICACION' not in df.columns:
                df['UBICACION'] = ''
            if 'CODIGO_SIIGO' not in df.columns:
                df['CODIGO_SIIGO'] = ''

            df['ORIGEN'] = origen
            self.logger.agregar_log(f"Archivo {archivo.name} procesado correctamente", 'exito')
            return df

        except Exception as e:
            self.logger.agregar_log(f"Error procesando {archivo.name}: {str(e)}", 'error')
            raise

    def leer_archivo_marca(self, archivo, marca):
        """Procesa archivos de marcas específicas (STIHL, SUZUKI, YAMAHA)."""
        mapeo_columnas = {
            'REFERENCIA': ['REFERENCIA', 'CODIGO', 'SKU', 'MODELO', 'PART NUMBER'],
            'DESCRIPCION': ['NOMBRE', 'DESCRIPCION', 'DESCRIPCIÓN', 'DESCRIP', 'PRODUCTO'],
            'CANTIDAD': ['CANTIDAD', 'QTY', 'QUANTITY', 'STOCK'],
            'UBICACION': ['UBICACION', 'UBICACIÓN', 'LOCALIZACION', 'ALMACEN']
        }
        columnas_requeridas = ['REFERENCIA', 'DESCRIPCION', 'CANTIDAD']

        df = self._procesar_archivo_generico(archivo, mapeo_columnas, columnas_requeridas, marca)
        return df[['REFERENCIA', 'DESCRIPCION', 'CANTIDAD', 'ORIGEN', 'UBICACION', 'CODIGO_SIIGO']]

    def leer_archivo_siigo(self, archivo):
        try:
            self.logger.agregar_log(f"Procesando archivo de valoración: {archivo.name}...")

            # Leer todo el archivo sin encabezados para buscar la fila correcta
            df_raw = pd.read_excel(archivo, header=None)

            # Buscar fila con encabezados
            fila_encabezados = None
            for i, fila in df_raw.iterrows():
                # Convertir todos los valores de la fila a string y limpiarlos
                valores = [str(cell).upper().strip() for cell in fila.values]

                # Verificar si encontramos los encabezados clave
                tiene_codigo = any(col in valores for col in ['CÓDIGO PRODUCTO', 'CODIGO PRODUCTO', 'CODIGO'])
                tiene_referencia = any(col in valores for col in ['REFERENCIA FÁBRICA', 'REFERENCIA FABRICA', 'REFERENCIA'])
                tiene_saldo = any(col in valores for col in ['SALDO CANTIDADES', 'SALDO', 'CANTIDAD'])

                if tiene_codigo and tiene_referencia and tiene_saldo:
                    fila_encabezados = i
                    self.logger.agregar_log(f"Encabezados encontrados en la fila {fila_encabezados + 1} (contando desde 1)")
                    break

            if fila_encabezados is None:
                # Mostrar las primeras filas para ayudar a diagnosticar
                muestra = df_raw.head(5).applymap(lambda x: str(x).upper().strip())
                self.logger.agregar_log(f"Primeras filas del archivo:\n{muestra}")
                raise ValueError("No se encontró una fila con todos los encabezados requeridos (Código Producto, Referencia Fábrica, Saldo Cantidades)")

            # Leer el archivo correctamente usando la fila de encabezados
            df = pd.read_excel(archivo, header=fila_encabezados)
            df.columns = df.columns.str.upper().str.strip()

            # Mapeo de columnas más flexible
            mapeo_columnas = {
                'CODIGO_PRODUCTO': ['CÓDIGO PRODUCTO', 'CODIGO PRODUCTO', 'CODIGO', 'CODIGO SIIGO'],
                'NOMBRE_PRODUCTO': ['NOMBRE PRODUCTO', 'DESCRIPCION', 'PRODUCTO', 'NOMBRE'],
                'REFERENCIA_FABRICA': ['REFERENCIA FÁBRICA', 'REFERENCIA FABRICA', 'REFERENCIA', 'SKU', 'MODELO'],
                'SALDO_CANTIDADES': ['SALDO CANTIDADES', 'SALDO', 'CANTIDAD', 'STOCK', 'EXISTENCIAS']
            }

            # Renombrar columnas según el mapeo
            columnas_renombradas = {}
            for col in df.columns:
                col_upper = str(col).upper().strip()
                for key, values in mapeo_columnas.items():
                    if any(v == col_upper for v in values):
                        columnas_renombradas[col] = key
                        break
                else:
                    columnas_renombradas[col] = col

            df = df.rename(columns=columnas_renombradas)

            # Verificar columnas requeridas
            columnas_requeridas = ['CODIGO_PRODUCTO', 'REFERENCIA_FABRICA', 'SALDO_CANTIDADES']
            for col in columnas_requeridas:
                if col not in df.columns:
                    # Mostrar las columnas disponibles para ayudar a diagnosticar
                    self.logger.agregar_log(f"Columnas disponibles: {list(df.columns)}")
                    raise ValueError(f"No se pudo identificar la columna '{col}'")

            # Si no existe NOMBRE_PRODUCTO, crearla vacía
            if 'NOMBRE_PRODUCTO' not in df.columns:
                df['NOMBRE_PRODUCTO'] = ''

            # Limpieza de datos
            df = df.dropna(how='all')  # Eliminar filas completamente vacías
            df['REFERENCIA_FABRICA'] = df['REFERENCIA_FABRICA'].astype(str).str.strip()
            df['NOMBRE_PRODUCTO'] = df['NOMBRE_PRODUCTO'].astype(str).str.strip()
            df['SALDO_CANTIDADES'] = pd.to_numeric(df['SALDO_CANTIDADES'], errors='coerce').fillna(0)

            # Extraer ubicación del nombre del producto
            df['UBICACION'] = df['NOMBRE_PRODUCTO'].str.extract(r'\((.*?)\)$').fillna('')
            df['DESCRIPCION'] = df['NOMBRE_PRODUCTO'].str.replace(r'\(.*?\)$', '', regex=True).str.strip()

            # Agregar información de origen
            df['ORIGEN'] = 'SIIGO'

            self.logger.agregar_log(f"Archivo de valoración procesado correctamente. Filas leídas: {len(df)}", 'exito')

            return df[['REFERENCIA_FABRICA', 'DESCRIPCION', 'SALDO_CANTIDADES', 'ORIGEN', 'UBICACION', 'CODIGO_PRODUCTO']].rename(columns={
                'REFERENCIA_FABRICA': 'REFERENCIA',
                'SALDO_CANTIDADES': 'CANTIDAD',
                'CODIGO_PRODUCTO': 'CODIGO_SIIGO'
            })

        except Exception as e:
            self.logger.agregar_log(f"Error procesando archivo de valoración: {str(e)}", 'error')
            raise

    def crear_consolidado(self):
        try:
            self.logger.agregar_log("Iniciando proceso de consolidación...")

            # Verificar carpeta 'inputs'
            inputs_dir = Path("inputs")
            if not inputs_dir.exists():
                raise FileNotFoundError("No se encontró la carpeta 'inputs'")

            # Buscar archivos requeridos
            archivos_requeridos = {
                'STIHL': None,
                'SUZUKI': None,
                'YAMAHA': None,
                'VALORACION': None
            }

            self.logger.agregar_log("Buscando archivos en la carpeta inputs...")
            for archivo in inputs_dir.glob("*.xlsx"):
                nombre = archivo.stem.upper()
                if 'STIHL' in nombre:
                    archivos_requeridos['STIHL'] = archivo
                elif 'SUZUKI' in nombre:
                    archivos_requeridos['SUZUKI'] = archivo
                elif 'YAMAHA' in nombre:
                    archivos_requeridos['YAMAHA'] = archivo
                elif 'VALORACION' in nombre or 'VALORACIÓN' in nombre:
                    archivos_requeridos['VALORACION'] = archivo

            # Verificar archivos faltantes
            for marca, archivo in archivos_requeridos.items():
                if archivo is None:
                    raise FileNotFoundError(f"No se encontró el archivo para {marca}")

            # Procesar cada archivo
            dfs = []
            for marca in ['STIHL', 'SUZUKI', 'YAMAHA']:
                df_marca = self.leer_archivo_marca(archivos_requeridos[marca], marca)
                # Asegurar columnas necesarias
                df_marca['STIHL'] = 0
                df_marca['SUZUKI'] = 0
                df_marca['YAMAHA'] = 0
                df_marca[marca] = df_marca['CANTIDAD']
                dfs.append(df_marca)

            df_siigo = self.leer_archivo_siigo(archivos_requeridos['VALORACION'])
            # Asegurar columnas necesarias
            df_siigo['STIHL'] = 0
            df_siigo['SUZUKI'] = 0
            df_siigo['YAMAHA'] = 0
            df_siigo['SIIGO'] = df_siigo['CANTIDAD']
            dfs.append(df_siigo)

            # Consolidar datos
            consolidado = pd.concat(dfs, ignore_index=True)

            # Crear carpeta 'outputs' si no existe
            output_dir = Path("outputs")
            output_dir.mkdir(exist_ok=True)

            # Guardar consolidado en archivo separado
            consolidado_file = output_dir / "Consolidado_Inventarios.xlsx"
            with pd.ExcelWriter(consolidado_file, engine='openpyxl') as writer:
                consolidado.to_excel(writer, sheet_name='Consolidado', index=False)
                writer.sheets['Consolidado'].sheet_state = 'visible'

            self.logger.agregar_log(f"Archivo consolidado creado: {consolidado_file}", 'exito')
            return consolidado_file

        except Exception as e:
            self.logger.agregar_log(f"Error durante la consolidación: {str(e)}", 'error')
            raise

    def procesar_consolidado(self, ruta_consolidado):
        """Versión idéntica a la original que funcionaba correctamente"""
        try:
            self.logger.agregar_log(f"\nProcesando archivo consolidado: {ruta_consolidado}...")

            # Leer datos consolidados (igual que antes)
            df = pd.read_excel(ruta_consolidado, sheet_name="Consolidado")
            df.columns = df.columns.str.strip()
            df["REFERENCIA"] = df["REFERENCIA"].astype(str).str.strip().replace("nan", "")

            # Separar referencias válidas vs inválidas (igual que antes)
            con_ref = df[df["REFERENCIA"] != ""].copy()
            sin_ref = df[df["REFERENCIA"] == ""].copy()

            # Procesar referencias válidas (método original)
            pivot = con_ref.pivot_table(
                index="REFERENCIA",
                columns="ORIGEN",
                values="CANTIDAD",
                aggfunc="sum",
                fill_value=0
            ).reset_index()

            # Asegurar columnas requeridas (igual que antes)
            for marca in ['STIHL', 'SUZUKI', 'YAMAHA', 'SIIGO']:
                if marca not in pivot.columns:
                    pivot[marca] = 0

            # Obtener descripciones y ubicaciones (método original)
            info_adicional = con_ref.groupby("REFERENCIA").agg({
                'DESCRIPCION': 'first',
                'UBICACION': lambda x: ', '.join(str(v) for v in set(x) if pd.notna(v) and v != ''),
                'CODIGO_SIIGO': 'first'
            }).reset_index()

            pivot = pivot.merge(info_adicional, on="REFERENCIA", how="left")
            pivot["DIFERENCIA"] = pivot[['STIHL', 'SUZUKI', 'YAMAHA']].sum(axis=1) - pivot["SIIGO"]

            # Procesar items sin referencia (método original)
            sin_ref["DIFERENCIA"] = sin_ref.apply(
                lambda x: x["CANTIDAD"] if x["ORIGEN"] in ['STIHL', 'SUZUKI', 'YAMAHA'] else -x["CANTIDAD"],
                axis=1
            )

            # Columnas comunes (método original)
            pivot["Inventario manual"] = pivot[["STIHL", "SUZUKI", "YAMAHA"]].sum(axis=1)
            sin_ref["Inventario manual"] = sin_ref[["STIHL", "SUZUKI", "YAMAHA"]].sum(axis=1)

            # Determinar MARCA (método original)
            def obtener_marca(row):
                marcas = [m for m in ["STIHL", "SUZUKI", "YAMAHA"] if row[m] > 0]
                return marcas[0] if len(marcas) == 1 else ("VARIAS" if len(marcas) > 1 else "")

            pivot["MARCA"] = pivot.apply(obtener_marca, axis=1)
            sin_ref["MARCA"] = sin_ref.apply(obtener_marca, axis=1)

            # Columnas finales (método original)
            columnas_finales = [
                "REFERENCIA", "STIHL", "SUZUKI", "YAMAHA", "Inventario manual", "SIIGO",
                "DESCRIPCION", "DIFERENCIA", "UBICACION", "CODIGO_SIIGO", "MARCA"
            ]

            tabla_final = pd.concat([
                pivot[columnas_finales],
                sin_ref[columnas_finales]
            ], ignore_index=True)

            # Crear archivo de análisis (método original)
            analisis_file = Path("outputs") / "Analisis_Comparativo.xlsx"
            with pd.ExcelWriter(analisis_file, engine='openpyxl') as writer:
                # Hoja Comparativo
                tabla_final.to_excel(writer, sheet_name='Comparativo', index=False)

                # Formatear diferencias
                ws = writer.sheets['Comparativo']
                rojo = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
                    for cell in row:
                        if cell.value != 0:
                            cell.fill = rojo

                # Hoja Resumen
                resumen = pd.DataFrame({
                    'Métrica': [
                        'Total ítems',
                        'Ítems con diferencias',
                        'Mayores en SIIGO',
                        'Mayores en Marcas',
                        'Diferencia total'
                    ],
                    'Valor': [
                        len(tabla_final),
                        sum(tabla_final['DIFERENCIA'] != 0),
                        sum(tabla_final['DIFERENCIA'] < 0),
                        sum(tabla_final['DIFERENCIA'] > 0),
                        tabla_final['DIFERENCIA'].sum()
                    ]
                })
                resumen.to_excel(writer, sheet_name='Resumen', index=False)

                # Asegurar visibilidad
                for sheet in writer.sheets.values():
                    sheet.sheet_state = 'visible'

            self.logger.agregar_log(f"Análisis comparativo creado: {analisis_file}", 'exito')
            return analisis_file

        except Exception as e:
            self.logger.agregar_log(f"Error durante el procesamiento: {str(e)}", 'error')
            raise

